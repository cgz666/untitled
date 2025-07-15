import requests
from openpyxl import Workbook
from openpyxl.styles import Font
import pythoncom
import win32com.client as win32
import os

class Device_management():
    def __init__(self):
        cookie_url = "http://clound.gxtower.cn:3980/tt/get_aiot_cookie?id=2"
        res = requests.get(cookie_url)
        cookie = res.text.strip()
        self.url = "https://zlzywg.chinatowercom.cn:8070/api/semp/sempDev/listSempDev"
        self.headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
        "Authorization": f"{cookie}",
        "Connection": "keep-alive",
        "Content-Length": "86",
        "Content-Type": "application/json;charset=UTF-8",
        "Cookie": "HWWAFSESID=2e34f79b8eced3aeab; HWWAFSESTIME=1750918105599",
        "Host": "zlzywg.chinatowercom.cn:8070",
        "Origin": "https://zlzywg.chinatowercom.cn:8070",
        "Referer": "https://zlzywg.chinatowercom.cn:8070/basicdata/terminal",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36 Edg/136.0.0.0",
        "sec-ch-ua": "\"Chromium\";v=\"136\", \"Microsoft Edge\";v=\"136\", \"Not.A/Brand\";v=\"99\"",
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": "\"Windows\""
}
        self.INDEX = os.getcwd()
        self.xls_path = os.path.join(self.INDEX, 'xls')
        self.output_path = os.path.join(self.INDEX, 'output')
        self.output_name = os.path.join(self.output_path, '设备信息-结果.xlsx')
        self.down_name = os.path.join(self.xls_path, '设备信息.xlsx')
        self.model_name = os.path.join(self.xls_path, '模板.xlsx')
    def down(self):
        page = 1
        page_size = 100
        all_devices = []
        # 定义表头和字段映射
        field_mapping = {
            '运行状态': 'onlineStatusName',
            '设备编码': 'devCode',
            '设备名称': 'devName',
            '站址编码': 'stationCode',
            '站址名称': 'stationName',
            '站址类型': 'stationTypeName',
            '站址来源': 'stationSourceCodeName',
            '站址地址': 'stationAddress',
            '资源编码': 'resCode',
            '业务类型': 'devBusinessTypeName',
            '资源类型': 'devResTypeName',
            '设备种类': 'devTypeName',
            '设备类型': 'devChildTypeName',
            '接入方式': 'accessWayName',
            '在服状态': 'serveStatusName',
            '监控方式': 'monitorName',
            '交维状态': 'maintainStatusName',
            '交维二级状态': 'maintainSecondName',
            '维护编码': 'maintainCode',
            '所属节点': 'colonyCode',
            '设备厂商': 'devFirmName',
            '设备型号': 'devModel',
            '产权归属': 'propertyOwnerName',
            '维护单位': 'maintainCompanyName',
            '管理区域（省）': 'admProvinceName',
            '管理区域(市)': 'admCityName',
            '管理区域(区)': 'admCountyName',
            '资源创建时间': 'resourceCreationTime',
            '实物确认时间（资源）': 'resAckTime',
            '入网时间（资源）': 'accessTime',
            '设备录入时间（运管）': 'recordTime',
            '创建时间（专业网管）': 'createTime',
            '更新时间（专业网管）': 'updateTime',
            '备电方式': 'powerSupplyWayName',
            '传输方式': 'transmissionMethodName',
            '用电方式': 'powerUseWayName',
            '设备挂高（米）': 'height',
            '设备重量（千克）': 'deviceWeightKg',
            '设备占用面积': 'deviceFootprint',
            '电压': 'voltage',
            '功率': 'power',
            '额定功率': 'ratedPower'
        }
        while True:
            data = {
                "admProvinceCode": "450000",
                "devClass": "0",
                "serveStatus": "1,3",
                "page": page,
                "pageSize": page_size
            }
            response = requests.post(url=self.url, headers=self.headers, json=data)
            response_json = response.json()
            current_page_data = response_json.get('data', {}).get('data', [])

            # 获取总页数信息（如果API返回）
            total = response_json.get('data', {}).get('total', 0)
            total_pages = (total + page_size - 1) // page_size
            if page > total_pages:
                break
            else:
                print(f"已获取第 {page}/{total_pages} 页，共 {len(current_page_data)} 条记录")

            # 处理当前页数据并添加到总列表
            for device in current_page_data:
                processed_device = {
                    chinese_header: device.get(json_key, '')
                    for chinese_header, json_key in field_mapping.items()
                }
                all_devices.append(processed_device)
            page += 1

        wb = Workbook()
        ws = wb.active
        ws.title = "设备数据"

        # 获取表头（中文列名）
        headers = list(field_mapping.keys())

        # 写入表头并设置字体为粗体
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)

        # 写入所有数据
        for row_idx, device in enumerate(all_devices, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx).value = device[header]

        # 保存Excel文件
        wb.save(self.down_name)
        print(f"数据已成功保存到 {self.down_name}，共 {len(all_devices)} 条记录")
        return True
    def excel_process(self):
        """
        处理Excel文件，将指定文件夹中的数据文件内容复制到主表文件中。

        :param index_path: 文件夹路径
        """
        print('1、把数据文件和通报模板放在同一文件夹下')
        print('2、打开上述文件，如果提示保护视图则取消（报错大概率是这个问题），如果提示别的东西请点击掉，保证程序能够编辑文档')

        # 初始化 COM 库
        pythoncom.CoInitialize()

        try:
            # 打开模板文件
            xl = win32.gencache.EnsureDispatch('Excel.Application')  # 开启excel软件
            xl.Visible = True  # 窗口是否可见
            xl.DisplayAlerts = False
            workbook_main = xl.Workbooks.Open(self.model_name)  # 打开上述路径文件
            #===================处理第一个子表======================
            # 打开下载文件
            workbook_data = xl.Workbooks.Open(self.down_name)
            sheet_data = workbook_data.Sheets('设备数据')
            sheet_main = workbook_main.Sheets('智联设备-项目名称')

            # 动态获取数据的实际范围
            last_row = sheet_data.Cells(sheet_data.Rows.Count, 1).End(win32.constants.xlUp).Row
            source_range = sheet_data.Range(f'A2:AP{last_row}')  # 从A2开始复制

            # 1. 只清除C-AR列的数据（不包含第一行表头）
            last_clear_row = sheet_main.UsedRange.Rows.Count
            if last_clear_row > 1:
                sheet_main.Range(f"C2:AR{last_clear_row}").ClearContents()

            # 只粘贴值到目标表
            source_range.Copy()
            sheet_main.Range('C2').PasteSpecial(Paste=win32.constants.xlPasteValues)
            xl.CutCopyMode = False  # 释放剪切板

            # 使用autofill续写A列数据
            sheet_main.Range('A2').AutoFill(sheet_main.Range(f'A2:A{last_row}'), win32.constants.xlFillDefault)

            #===================处理第二个子表======================
            # 获取main_sheet的第二个子表
            sheet_main_2 = workbook_main.Sheets('剔除虚拟站、社会站项目站址')  # 替换为实际的子表名称

            last_clear_row_2 = sheet_main_2.UsedRange.Rows.Count
            if last_clear_row_2 > 1:
                sheet_main_2.Range(f"D2:AS{last_clear_row_2}").ClearContents()

            # 只粘贴值到目标表
            source_range.Copy()
            sheet_main_2.Range('D2').PasteSpecial(Paste=win32.constants.xlPasteValues)
            xl.CutCopyMode = False  # 释放剪切板

            # 删除站址类型(H列)为‘虚拟站’的数据
            for row in range(last_row, 1, -1):  # 从最后一行向上遍历
                cell_value = sheet_main_2.Cells(row, 9).Value  # H列的值
                if cell_value == '虚拟站':
                    sheet_main_2.Rows(row).Delete()

            # 删除站址来源(I列)‘自有站’以外的数据
            for row in range(last_row, 1, -1):  # 从最后一行向上遍历
                cell_value = sheet_main_2.Cells(row, 10).Value  # I列的值
                if cell_value != '自有站':
                    sheet_main_2.Rows(row).Delete()

            # 使用autofill续写第二个子表的A列和B列数据
            sheet_main_2.Range('A2').AutoFill(sheet_main_2.Range(f'A2:A{last_row}'), win32.constants.xlFillDefault)
            sheet_main_2.Range('B2').AutoFill(sheet_main_2.Range(f'B2:B{last_row}'), win32.constants.xlFillDefault)

            #===================处理第三个子表======================
            # 获取main_sheet的第三个子表
            sheet_main_3 = workbook_main.Sheets('已交维剔除虚拟站、社会站项目站址')  # 替换为实际的子表名称

            last_clear_row_3 = sheet_main_3.UsedRange.Rows.Count
            if last_clear_row_3 > 1:
                sheet_main_3.Range(f"D2:AS{last_clear_row_3}").ClearContents()

            # 只粘贴值到目标表
            source_range.Copy()
            sheet_main_3.Range('D2').PasteSpecial(Paste=win32.constants.xlPasteValues)
            xl.CutCopyMode = False  # 释放剪切板

            # 删除站址类型(H列)为‘虚拟站’的数据
            for row in range(last_row, 1, -1):  # 从最后一行向上遍历
                cell_value = sheet_main_3.Cells(row, 9).Value  # H列的值
                if cell_value == '虚拟站':
                    sheet_main_3.Rows(row).Delete()

            # 删除站址来源(I列)‘自有站’以外的数据
            for row in range(last_row, 1, -1):  # 从最后一行向上遍历
                cell_value = sheet_main_3.Cells(row, 10).Value  # I列的值
                if cell_value != '自有站':
                    sheet_main_3.Rows(row).Delete()

            # 删除‘交维状态’(T列)为‘已交维’和‘交维待补码’以外的数据
            for row in range(last_row, 1, -1):  # 从最后一行向上遍历
                cell_value = sheet_main_3.Cells(row, 20).Value  # T列是第20列
                if cell_value not in ['已交维', '交维待补码']:
                    sheet_main_3.Rows(row).Delete()

            # 使用autofill续写第二个子表的A列和B列数据
            sheet_main_3.Range('A2').AutoFill(sheet_main_3.Range(f'A2:A{last_row}'), win32.constants.xlFillDefault)
            sheet_main_3.Range('B2').AutoFill(sheet_main_3.Range(f'B2:B{last_row}'), win32.constants.xlFillDefault)
            workbook_main.SaveAs(self.output_name)
            workbook_main.Close()
            xl.Quit()  # 关闭Excel应用程序
            print('已全部完成')
        except Exception as e:
            raise
        finally:
            # 释放 COM 库
            pythoncom.CoUninitialize()

    def main(self):
        self.down()
        self.excel_process()
if __name__ == "__main__":
    Device_management().main()
