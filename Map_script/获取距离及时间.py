from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import time
import openpyxl
import os
import random
import logging

# 配置日志记录
logging.basicConfig(
    filename='map_script.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

class MapScriptOptimized():
    def __init__(self):
        """初始化处理器"""
        self.INDEX = os.getcwd()
        self.output_path = os.path.join(self.INDEX, "output")
        self.driver_path = os.path.join(self.INDEX, "chromedriver.exe")
        self.chrome_path = os.path.join(self.INDEX, "chrome-win64/chrome.exe")

        # 文件路径配置
        self.save_path = os.path.join(self.INDEX, "xls")
        self.output_path = os.path.join(self.INDEX, "output")
        self.output_name = os.path.join(self.output_path, "站点信息-结果.xlsx")
        self.site_data = os.path.join(self.save_path, "站点信息.xlsx")

        # 初始化浏览器驱动
        self.driver = self.init_driver()
        self.wait = WebDriverWait(self.driver, 15)

        # 重试配置
        self.max_retries = 1  # 最大重试次数
        self.base_delay = 2  # 基础延迟秒数

    def init_driver(self):
        """初始化浏览器驱动，屏蔽控制台日志"""
        service = Service(executable_path=self.driver_path)
        option = webdriver.ChromeOptions()
        option.binary_location = self.chrome_path
        option.add_argument("--headless")
        option.add_argument("--disable-gpu")
        option.add_argument("--window-size=1920,1080")
        option.add_argument("--no-sandbox")
        option.add_argument("--disable-dev-shm-usage")
        option.add_argument("--blink-settings=imagesEnabled=false")  # 禁用图片加载
        option.add_argument("--disable-extensions")  # 禁用扩展

        # 屏蔽控制台日志
        option.add_experimental_option('excludeSwitches', ['enable-logging'])

        # 禁用不必要的浏览器功能
        option.add_argument("--disable-web-security")
        option.add_argument("--disable-notifications")
        option.add_argument("--disable-popup-blocking")

        return webdriver.Chrome(service=service, options=option)

    def get_driving_info(self, from_lng, from_lat, to_lng, to_lat):
        """获取两点间的驾驶距离和时间，带重试机制"""
        retries = 0
        while retries < self.max_retries:
            try:
                # 构建URL
                url = f'https://ditu.amap.com/dir?type=car&from%5Blnglat%5D={from_lng}%2C{from_lat}&from%5Bname%5D=%E8%B5%B7%E7%82%B9&to%5Blnglat%5D={to_lng}%2C{to_lat}&to%5Bname%5D=%E7%BB%88%E7%82%B9&src=uriapi&innersrc=uriapi&policy=1'
                self.driver.get(url)

                print(f"正在获取从 {from_lng},{from_lat} 到 {to_lng},{to_lat} 的路线信息 (尝试 {retries + 1}/{self.max_retries})...")

                # 使用显式等待替代固定等待
                element = self.wait.until(
                    EC.presence_of_element_located((By.CLASS_NAME, "dir_base_info"))
                )

                # 获取页面源码
                html = self.driver.page_source
                soup = BeautifulSoup(html, 'html.parser')

                # 查找方案一的距离和时间信息
                dir_base_info = soup.find('p', class_='dir_base_info')
                if dir_base_info:
                    spans = dir_base_info.find_all('span', class_='dir_base')
                    if len(spans) >= 2:
                        time_info = spans[0].text.strip()
                        distance_info = spans[1].text.strip()
                        print(f"成功获取路线信息: {time_info}, {distance_info}")
                        return time_info, distance_info

                print(f"尝试 {retries + 1}/{self.max_retries}: 未找到路线信息，尝试刷新...")
                self.driver.refresh()

            except Exception as e:
                print(f"尝试 {retries + 1}/{self.max_retries}: 获取路线信息时出错: {e}")
                logging.error(
                    f"尝试 {retries + 1}/{self.max_retries}: 获取路线信息时出错 - {from_lng},{from_lat} 到 {to_lng},{to_lat}: {str(e)}")

            # 指数退避策略：随重试次数增加延迟时间，并添加随机抖动
            delay = self.base_delay * (2 ** retries) + random.uniform(0, 1)
            print(f"等待 {delay:.2f} 秒后重试...")
            time.sleep(delay)
            retries += 1

        # 所有重试都失败
        print(f"警告: 从 {from_lng},{from_lat} 到 {to_lng},{to_lat} 的路线信息获取失败，已达到最大重试次数")
        logging.warning(f"获取失败: {from_lng},{from_lat} 到 {to_lng},{to_lat}")
        return None, None

    def process_row(self, ws, row, from_lng_col, from_lat_col, to_lng_col, to_lat_col, time_col, distance_col):
        """处理单行数据，带重试机制"""
        # 获取经纬度信息
        from_lng = ws.cell(row=row, column=from_lng_col).value
        from_lat = ws.cell(row=row, column=from_lat_col).value
        to_lng = ws.cell(row=row, column=to_lng_col).value
        to_lat = ws.cell(row=row, column=to_lat_col).value

        # 检查经纬度是否存在
        if not all([from_lng, from_lat, to_lng, to_lat]):
            print(f"第 {row} 行经纬度信息不完整，跳过")
            return False

        # 尝试获取数据，直到成功
        time_info, distance_info = self.get_driving_info(from_lng, from_lat, to_lng, to_lat)

        # 如果重试后仍失败，返回False
        if not time_info or not distance_info:
            print(f"第 {row} 行获取数据失败")
            return False

        # 写入结果
        ws.cell(row=row, column=time_col).value = time_info
        ws.cell(row=row, column=distance_col).value = distance_info

        print(f"第 {row} 行处理完成")
        return True

    def process_excel(self):
        """处理Excel文件，获取经纬度并写入距离和时间"""
        # 读取Excel文件
        wb = openpyxl.load_workbook(self.site_data)
        ws = wb.active

        # 确定列索引 (R=18, S=19, V=22, W=23, X=24, Y=25)
        from_lng_col = 18  # R列
        from_lat_col = 19  # S列
        to_lng_col = 22  # V列
        to_lat_col = 23  # W列
        time_col = 24  # X列
        distance_col = 25  # Y列

        # 添加表头
        ws.cell(row=1, column=time_col).value = "驾驶时间"
        ws.cell(row=1, column=distance_col).value = "驾驶距离"

        # 获取最大行数
        max_row = ws.max_row

        # 顺序处理每一行数据
        completed_count = 0
        for row in range(2, max_row + 1):
            try:
                success = self.process_row(
                    ws, row,
                    from_lng_col, from_lat_col,
                    to_lng_col, to_lat_col,
                    time_col, distance_col
                )

                if success:
                    completed_count += 1

                # 每处理完10行保存一次
                if row % 10 == 0:
                    wb.save(self.output_name)
                    print(f"已保存进度，当前完成: {completed_count}/{max_row - 1} 行")

            except Exception as e:
                print(f"处理第 {row} 行时出错: {e}")
                logging.error(f"处理第 {row} 行时出错: {str(e)}")

        # 最终保存
        wb.save(self.output_name)
        print(f"处理完成，结果已保存到: {self.output_name}")
        print(f"成功处理: {completed_count}/{max_row - 1} 行")
        if completed_count < max_row - 1:
            print(f"警告: {max_row - 1 - completed_count} 行未能成功获取数据，请查看日志文件获取详情")

    def main(self):
        try:
            print("开始处理地图数据...")
            self.process_excel()
        finally:
            self.driver.quit()
            print("程序已结束")


if __name__ == "__main__":
    MapScriptOptimized().main()