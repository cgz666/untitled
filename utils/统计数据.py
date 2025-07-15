import openpyxl
from openpyxl import Workbook
import os
import sys

index = os.getcwd()
model_file = os.path.join(index, 'data.xlsx')
# 打开Excel文件
wb = openpyxl.load_workbook(model_file)

# 获取工作表
sheet_current = wb['整流模块电流']
sheet_temp = wb['整流模块温度']
sheet3 = wb['Sheet3']

# 初始化sheet3的标题行
if sheet3.max_row < 1:
    sheet3['A1'] = '站址'
    sheet3['B1'] = '计数'

# 用于存储站址和计数的字典
site_count = {}

# 辅助函数：将值转换为数值类型
def convert_to_number(value):
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        raise

# 辅助函数：显示进度条
def show_progress(current, total, prefix="处理中", suffix="完成", length=50):
    percent = f"{100 * (current / float(total)):.1f}"
    filled_length = int(length * current // total)
    bar = '█' * filled_length + '-' * (length - filled_length)
    sys.stdout.write(f'\r{prefix} |{bar}| {percent}% {suffix}')
    sys.stdout.flush()
    if current == total:
        print()  # 处理完成后换行

# 处理整流模块电流工作表
print("开始处理整流模块电流数据...")
total_current_rows = sheet_current.max_row - 1  # 减去标题行
for i, row in enumerate(range(2, sheet_current.max_row + 1)):
    site = sheet_current.cell(row=row, column=4).value  # D列为站址
    value = sheet_current.cell(row=row, column=14).value  # N列为实测值
    value_num = convert_to_number(value)
    if value_num is not None and 0 < value_num < 70:
        if site not in site_count:
            site_count[site] = 1
        else:
            site_count[site] += 1
    show_progress(i + 1, total_current_rows, "电流数据处理进度")

# 处理整流模块温度工作表
print("\n开始处理整流模块温度数据...")
total_temp_rows = sheet_temp.max_row - 1  # 减去标题行
for i, row in enumerate(range(2, sheet_temp.max_row + 1)):
    site = sheet_temp.cell(row=row, column=4).value  # D列为站址
    value = sheet_temp.cell(row=row, column=14).value  # N列为实测值
    value_num = convert_to_number(value)
    if value_num is not None and 0 < value_num < 100:
        if site not in site_count:
            site_count[site] = 1
        else:
            site_count[site] += 1
    show_progress(i + 1, total_temp_rows, "温度数据处理进度")

# 将数据写入sheet3
print("\n开始写入统计结果...")
total_sites = len(site_count)
for i, (site, count) in enumerate(site_count.items()):
    found = False
    for row in range(2, sheet3.max_row + 1):
        if sheet3.cell(row=row, column=1).value == site:
            sheet3.cell(row=row, column=2).value += count
            found = True
            break
    if not found:
        sheet3.append([site, count])
    show_progress(i + 1, total_sites, "结果写入进度")

# 保存Excel文件
wb.save('output.xlsx')
print("\n数据处理完成，结果已保存至 output.xlsx")