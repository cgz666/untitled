import requests
import pythoncom
import time
import openpyxl
import win32com.client as win32
from openpyxl.styles import Font
import os
from datetime import datetime, timedelta

class Average_online_rate():
    def __init__(self):
        cookie_url = "http://clound.gxtower.cn:3980/tt/get_aiot_cookie?id=2"
        res = requests.get(cookie_url)
        if res.status_code == 200:
            cookie = res.text.strip()  # 获取响应内容并去除两端的空白字符
        else:
            print(f"Failed to get cookie: {res.status_code}")
            cookie = None
        self.url = "https://zlzywg.chinatowercom.cn:8070/api/sa/deviceAvgOnlineRateStatistics/deviceAvgOnlineRateAreaList"
        self.headers = {
            "Accept": "application/json, text/plain, */*",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
            "Authorization": f"{cookie}",
            "Connection": "keep-alive",
            "Content-Length": "122",
            "Content-Type": "application/json;charset=UTF-8",
            "Cookie": "HWWAFSESID=b326e738d641280da8; HWWAFSESTIME=1750062702120",
            "Host": "zlzywg.chinatowercom.cn:8070",
            "Origin": "https://zlzywg.chinatowercom.cn:8070",
            "Referer": "https://zlzywg.chinatowercom.cn:8070/statAnalysis/onRate",
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36 Edg/136.0.0.0",
            "sec-ch-ua": '"Chromium";v="136", "Microsoft Edge";v="136", "Not.A/Brand";v="99"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": "Windows"
        }
        INDEX = os.getcwd()
        self.AREA_CODE_MAPPING = {
            "450100": "南宁",
            "450200": "柳州",
            "450300": "桂林",
            "450400": "梧州",
            "450500": "北海",
            "450600": "防城港",
            "450700": "钦州",
            "450800": "贵港",
            "450900": "玉林",
            "451000": "百色",
            "451100": "贺州",
            "451200": "河池",
            "451300": "来宾",
            "451400": "崇左",
            "450000": "广西"
        }
        # 获取当前日期和前一天日期
        self.today = datetime.now()
        self.yesterday = self.today - timedelta(days=1)

        # 根据日期动态确定数据范围
        if self.today.day == 1:
            # 1号：获取上个月全量数据
            self.last_month = self.yesterday.month
            self.last_month_year = self.yesterday.year
            self.start_date = datetime(self.last_month_year, self.last_month, 1)
            self.date_list = [(self.start_date + timedelta(days=i)).strftime("%Y-%m-%d")
                              for i in range((self.yesterday - self.start_date).days + 1)]
        elif self.today.day == 2:
            # 2号：获取当月1号数据
            self.start_date = datetime(self.today.year, self.today.month, 1)
            self.date_list = [self.start_date.strftime("%Y-%m-%d")]
        else:
            # 其他日期：获取当月1号到昨天的数据
            self.start_date = datetime(self.today.year, self.today.month, 1)
            self.date_list = [(self.start_date + timedelta(days=i)).strftime("%Y-%m-%d")
                              for i in range((self.yesterday - self.start_date).days + 1)]

        self.target_area_codes = list(self.AREA_CODE_MAPPING.keys())
        self.date_list = [(self.start_date + timedelta(days=i)).strftime("%Y-%m-%d") for i in range((self.yesterday - self.start_date).days + 1)]
        self.save_path = os.path.join(INDEX, "xls")
        self.output_path = os.path.join(INDEX, "output")
        self.output_name = os.path.join(self.output_path, "智联设备平均在线率统计-结果.xlsx")
        self.model_path = os.path.join(self.save_path, "模板.xlsx")
        self.filename = os.path.join(self.save_path, "data.xlsx")

    def get_data(self):
        info_list = []

        # 遍历每个日期
        for date in self.date_list:
            data = {
                "admProvinceCode": "450000",
                "statisticsTimeFrom": date,
                "statisticsTimeTo": date,
                "onlineRateIncludeCustom": 0
            }
            try:
                response = requests.post(url=self.url, headers=self.headers, json=data)
                if response.status_code == 200:
                    response_json = response.json()
                    if 'data' in response_json and 'list' in response_json['data']:
                        # 处理所有数据项，包括省级和地市
                        for item in response_json['data']['list']:
                            area_code = item.get('areaCode', '')

                            if area_code in self.target_area_codes:
                                # 格式化日期并添加到数据项中
                                date_obj = datetime.strptime(date, "%Y-%m-%d")
                                chinese_date = f"{date_obj.month}月{date_obj.day}日"
                                item['query_date'] = chinese_date
                                item['short_city_name'] = self.AREA_CODE_MAPPING.get(area_code, area_code)

                                processed_item = {
                                    'areaCode': area_code,
                                    'short_city_name': item['short_city_name'],
                                    'query_date': item['query_date'],
                                    'onlineRate': item.get('onlineRate', ''),
                                    'monitoredOnlineRate': item.get('monitoredOnlineRate', ''),
                                    'localAreaOnlineRate': item.get('localAreaOnlineRate', ''),
                                    'whiteListOnlineRate': item.get('whiteListOnlineRate', ''),
                                }
                                info_list.append(processed_item)

                            if 'children' in item:
                                for child in item['children']:
                                    if child.get('areaCode') in self.target_area_codes:
                                        # 格式化日期并添加到数据项中
                                        date_obj = datetime.strptime(date, "%Y-%m-%d")
                                        chinese_date = f"{date_obj.month}月{date_obj.day}日"
                                        child['query_date'] = chinese_date
                                        child['short_city_name'] = self.AREA_CODE_MAPPING.get(
                                            child.get('areaCode', ''),
                                            child.get('areaCode', '')
                                        )
                                        info_list.append(child)
                        print(f"成功获取 {date} 的数据")
            except Exception:
                raise
            time.sleep(1)  # 避免过于频繁的请求

        # 按地市简称和日期排序数据
        info_list.sort(key=lambda x: (list(self.AREA_CODE_MAPPING.keys()).index(x.get('areaCode', '')),x.get('statisticsTimeFrom', '')))
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "设备在线率统计"

            # 设置表头
            headers = ['管理区域（市）', '日期', '平均在线率',
                       '总部可监控平均在线率', '省市监控平均在线率', '白名单平均在线率']

            # 写入表头并设置字体加粗
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)

            # 写入数据
            for row_idx, item in enumerate(info_list, 2):  # 从第2行开始
                ws.cell(row=row_idx, column=1, value=item.get('short_city_name', ''))
                ws.cell(row=row_idx, column=2, value=item.get('query_date', ''))
                ws.cell(row=row_idx, column=3, value=item.get('onlineRate', ''))
                ws.cell(row=row_idx, column=4, value=item.get('monitoredOnlineRate', ''))
                ws.cell(row=row_idx, column=5, value=item.get('localAreaOnlineRate', ''))
                ws.cell(row=row_idx, column=6, value=item.get('whiteListOnlineRate', ''))

            # 调整列宽以适应内容
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length * 1.2

            wb.save(self.filename)

            print(f'文件已保存到{self.filename}')
        except Exception :
            raise

    def excel_process(self):
        """
        处理Excel文件，将指定文件夹中的数据文件内容复制到主表文件中。

        :param index_path: 文件夹路径
        """
        print('1、把数据文件和通报模板放在同一文件夹下')
        print('2、打开上述文件，如果提示保护视图则取消（报错大概率是这个问题），如果提示别的东西请点击掉，保证程序能够编辑文档')

        # 初始化 COM 库
        pythoncom.CoInitialize()

        try:
            # 打开模板文件
            xl = win32.gencache.EnsureDispatch('Excel.Application')  # 开启excel软件
            xl.DisplayAlerts = False
            xl.Visible = True  # 窗口是否可见
            workbook_main = xl.Workbooks.Open(self.model_path)  # 打开上述路径文件

            # 打开下载文件
            workbook_data = xl.Workbooks.Open(self.filename)
            sheet_data = workbook_data.Sheets('设备在线率统计')
            sheet_main = workbook_main.Sheets('智联设备平均在线率统计')

            # 动态获取数据的实际范围
            last_row = sheet_data.Cells(sheet_data.Rows.Count, 1).End(win32.constants.xlUp).Row
            source_range = sheet_data.Range(f'A2:F{last_row}')  # 从A2开始复制

            # 1. 只清除A-F列的数据（不包含第一行表头）
            last_clear_row = sheet_main.UsedRange.Rows.Count
            if last_clear_row > 1:
                sheet_main.Range(f"A2:F{last_clear_row}").ClearContents()

            # 只粘贴值到目标表
            source_range.Copy()
            sheet_main.Range('A2').PasteSpecial(Paste=win32.constants.xlPasteValues)
            xl.CutCopyMode = False  # 释放剪切板

            # 获取模板的格式行（通常是第二行，假设表头下一行是格式样本）
            format_row = 2  # 模板中的格式行
            format_range = sheet_main.Range(f"A{format_row}:F{format_row}")

            # 获取新数据的范围（从第2行到最后一行）
            new_data_range = sheet_main.Range(f"A2:F{last_row}")

            # 应用模板格式到新数据
            format_range.Copy()
            new_data_range.PasteSpecial(Paste=win32.constants.xlPasteFormats)
            xl.CutCopyMode = False  # 释放剪切板

            # ===== 为每个地市生成公式 =====
            # 定义地市列表和对应的列
            cities = list(self.AREA_CODE_MAPPING.values())[:-1]  # 排除广西
            rate_columns = {'平均在线率': 'I', '总部可监控平均在线率': 'J', '省市监控平均在线率': 'K', '白名单平均在线率': 'L'}

            # 获取表头数据
            headers = [sheet_data.Cells(1, col_idx).Value for col_idx in range(1, 7)]

            # 获取数据区域的最后一行
            data_last_row = sheet_main.Cells(sheet_main.Rows.Count, 1).End(win32.constants.xlUp).Row

            # 遍历每个地市和每个在线率类型
            for city_idx, city in enumerate(cities, start=2):  # 从H2开始
                for rate_name, col_letter in rate_columns.items():
                    # 获取数据列索引
                    data_col_letter = chr(ord('A') + headers.index(rate_name))

                    # 构建公式：=AVERAGEIF(数据列, 地市列, 地市名称)
                    formula = f'=AVERAGEIF($A$2:$A${data_last_row},"{city}",{data_col_letter}$2:{data_col_letter}${data_last_row})'

                    # 设置公式到对应单元格
                    sheet_main.Range(f"{col_letter}{city_idx}").Formula = formula

            # ===== 新增：将C到F列文本型数字转换为数字 =====
            # 获取C到F列的数据范围
            convert_range = f"C2:F{data_last_row}"
            # 使用Value = Value技巧将文本型数字转换为数值型
            sheet_main.Range(convert_range).Value = sheet_main.Range(convert_range).Value

            workbook_data.Close(SaveChanges=False)
            workbook_main.SaveAs(self.output_name)
            workbook_main.Close()
            xl.Quit()
            print('已全部完成')
        except Exception as e:
            raise
        finally:
            # 释放 COM 库
            pythoncom.CoUninitialize()

    def main(self):
        self.get_data()
        self.excel_process()
if __name__ == "__main__":
    Average_online_rate().main()